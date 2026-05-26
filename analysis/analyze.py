"""Cognitive Battery - Analysis Template (Python).

Reads Excel or JSON exports from Cognitive Battery, applies pre-registered outlier
thresholds recorded in each file, and writes a participant-level summary CSV.

Usage:
    python analyze.py <export-directory> [output.csv]

Requires: Python 3.9+. Ex-Gaussian fitting additionally requires scipy
and numpy — if not installed, those columns are omitted with a warning.
Excel input requires openpyxl.
"""

from __future__ import annotations

import csv
import hashlib
import json
import sys
from pathlib import Path
from statistics import mean, median
from typing import Any, Iterable

try:
    import numpy as np
    from scipy import optimize, special
    _SCIPY_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    _SCIPY_AVAILABLE = False

try:
    from openpyxl import load_workbook
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    _OPENPYXL_AVAILABLE = False


def fit_ex_gaussian(rts: Iterable[float]) -> dict | None:
    """Fit ex-Gaussian (mu, sigma, tau) to response times via MLE.

    Returns {'ex_mu', 'ex_sigma', 'ex_tau', 'ex_converged'} or None if scipy
    is unavailable or the sample is too small (<15 RTs).
    """
    if not _SCIPY_AVAILABLE:
        return None
    rts = np.asarray(list(rts), dtype=float)
    rts = rts[np.isfinite(rts) & (rts > 0)]
    if len(rts) < 15:
        return None

    def neg_log_likelihood(params):
        mu, sigma, tau = params
        if sigma <= 0 or tau <= 0:
            return np.inf
        arg = (mu + (sigma ** 2) / tau - rts) / (np.sqrt(2) * sigma)
        log_inner = (mu - rts) / tau + (sigma ** 2) / (2 * tau ** 2)
        erfc_vals = special.erfc(arg)
        # Avoid log(0): clip to tiny positive
        inner = np.exp(log_inner) * erfc_vals
        inner = np.clip(inner, 1e-300, None)
        return -np.sum(np.log(inner / (2 * tau)))

    rt_mean = float(np.mean(rts))
    rt_std = float(np.std(rts))
    initial = [rt_mean - rt_std * 0.5, max(rt_std * 0.5, 1.0), max(rt_std * 0.5, 1.0)]
    result = optimize.minimize(neg_log_likelihood, initial, method="Nelder-Mead",
                               options={"maxiter": 2000, "xatol": 1e-4, "fatol": 1e-4})
    mu, sigma, tau = result.x
    return {
        "ex_mu": float(mu),
        "ex_sigma": float(sigma),
        "ex_tau": float(tau),
        "ex_converged": bool(result.success),
    }


def get(d: dict | None, *path, default=None):
    cur: Any = d
    for key in path:
        if not isinstance(cur, dict) or key not in cur:
            return default
        cur = cur[key]
    return cur if cur is not None else default


def summarise_task(trials: list[dict] | None, rt_low: float, rt_high: float) -> dict:
    trials = trials or []
    if not trials:
        return {"n_trials": 0, "n_kept_rt": 0}
    rts = [t.get("rt") for t in trials if isinstance(t.get("rt"), (int, float))]
    kept = [t for t in trials if isinstance(t.get("rt"), (int, float)) and rt_low <= t["rt"] <= rt_high]
    kept_correct = [t for t in kept if t.get("correct") == 1]
    return {
        "n_trials": len(trials),
        "n_kept_rt": len(kept),
        "accuracy_all": (sum(1 for t in trials if t.get("correct") == 1) / len(trials)) if trials else None,
        "accuracy_after_rt_exclude": (sum(1 for t in kept if t.get("correct") == 1) / len(kept)) if kept else None,
        "mean_rt_correct": mean(t["rt"] for t in kept_correct) if kept_correct else None,
        "median_rt_correct": median(t["rt"] for t in kept_correct) if kept_correct else None,
    }


def verify_integrity(payload: dict, recorded: str | None) -> bool | None:
    """Re-hash everything except the integrity field and compare.

    The browser computes the hash over `JSON.stringify(payload)` where `payload`
    already has had `integrity_sha256` stripped. Because JSON.stringify in the
    browser is not key-sorted, the Python comparison is only strict if we
    preserve key order. We therefore recompute from the on-disk file text after
    removing the last occurrence of the integrity field.
    """
    if not recorded:
        return None
    payload_no_hash = {k: v for k, v in payload.items() if k != "integrity_sha256"}
    # JS JSON.stringify default: no spaces, no key sorting. Python default
    # matches when we use separators=(",", ":") and preserve insertion order.
    reconstructed = json.dumps(payload_no_hash, ensure_ascii=False, separators=(",", ":"))
    recomputed = hashlib.sha256(reconstructed.encode("utf-8")).hexdigest()
    return recomputed == recorded


def qc_multiverse_rows(payload: dict) -> list[dict]:
    qc = payload.get("qc_multiverse", {})
    if isinstance(qc, dict):
        rows = qc.get("universes", [])
    else:
        rows = qc
    return rows if isinstance(rows, list) else []


def summarize_qc_multiverse(rows: list[dict]) -> dict:
    include_ids = [row.get("universe_id") for row in rows if row.get("include_candidate") == 1]
    exclude_ids = [row.get("universe_id") for row in rows if row.get("exclude_candidate") == 1]
    return {
        "qc_universe_count": len(rows),
        "qc_include_candidate_universes": ",".join(str(x) for x in include_ids if x),
        "qc_exclude_candidate_universes": ",".join(str(x) for x in exclude_ids if x),
        "qc_exclude_candidate_count": len(exclude_ids),
    }


def first_row(rows: list[dict]) -> dict:
    return rows[0] if rows else {}


def metadata_rows_to_dict(rows: list[dict]) -> dict:
    result = {}
    for row in rows:
        field = row.get("field")
        if field:
            result[str(field)] = row.get("value")
    return result


def sheet_to_records(workbook, sheet_name: str) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for header, value in zip(headers, row):
            if not header:
                continue
            record[header] = value
        if any(value is not None for value in record.values()):
            records.append(record)
    return records


def load_payload_from_xlsx(path: Path) -> dict:
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("Excel input requires openpyxl. Install it with: python -m pip install openpyxl")

    workbook = load_workbook(path, read_only=True, data_only=True)
    participant_row = first_row(sheet_to_records(workbook, "Participant"))
    protocol = metadata_rows_to_dict(sheet_to_records(workbook, "Protocol Metadata"))
    manifest = metadata_rows_to_dict(sheet_to_records(workbook, "Export Manifest"))
    session_quality = first_row(sheet_to_records(workbook, "Session Quality"))
    review_row = first_row(sheet_to_records(workbook, "Researcher Review"))
    research_metrics = first_row(sheet_to_records(workbook, "Research Metrics"))
    qc_rows = sheet_to_records(workbook, "QC Multiverse")
    task_metrics_long = sheet_to_records(workbook, "Task Metrics Long")
    session_events = sheet_to_records(workbook, "Session Events")

    raw_sheet_map = {
        "flanker_raw": "flanker",
        "dccs_raw": "dccs",
        "pattern_comparison_raw": "pattern-comparison",
        "list_sorting_raw": "list-sorting",
        "picture_sequence_raw": "picture-sequence",
    }
    trials = {
        test_id: sheet_to_records(workbook, sheet_name)
        for sheet_name, test_id in raw_sheet_map.items()
        if sheet_name in workbook.sheetnames
    }

    participant = {
        "participantId": participant_row.get("participantId") or manifest.get("participant_id"),
        "session_number": participant_row.get("session_number") or manifest.get("session_number"),
        "counterbalance_group": participant_row.get("counterbalance_group") or manifest.get("counterbalance_group"),
        "consent_version": participant_row.get("consent_version") or protocol.get("consent_version"),
        "age": participant_row.get("age"),
        "random_seed": participant_row.get("random_seed") or manifest.get("random_seed"),
        "viewing_distance_cm": participant_row.get("viewing_distance_cm"),
        "grayscale_confirmed": participant_row.get("grayscale_confirmed"),
    }
    environment = {
        "browser": participant_row.get("browser"),
        "viewportWidth": participant_row.get("viewport_width"),
        "viewportHeight": participant_row.get("viewport_height"),
        "devicePixelRatio": participant_row.get("device_pixel_ratio"),
        "refreshRateHzEstimate": participant_row.get("refresh_rate_hz_estimate"),
    }
    outlier_thresholds = {
        "rt_exclude_below_ms": participant_row.get("rt_exclude_below_ms"),
        "rt_too_slow_ms": participant_row.get("rt_too_slow_ms"),
        "rt_too_fast_ms": participant_row.get("rt_too_fast_ms"),
    }

    return {
        "version": protocol.get("app_version") or manifest.get("app_version"),
        "export_manifest": manifest,
        "participant": participant,
        "protocol": protocol,
        "outlier_thresholds": outlier_thresholds,
        "session_quality": session_quality,
        "quality_flags": review_row or session_quality,
        "environment": environment,
        "research_metrics": research_metrics,
        "task_metrics_long": task_metrics_long,
        "session_events": session_events,
        "trials": trials,
        "qc_multiverse": {
            "version": protocol.get("qc_multiverse_version") or manifest.get("qc_multiverse_version"),
            "universes": qc_rows,
        },
    }


def load_payload(path: Path) -> tuple[dict, bool | None]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        return payload, verify_integrity(payload, payload.get("integrity_sha256"))
    if suffix == ".xlsx":
        return load_payload_from_xlsx(path), None
    raise ValueError(f"Unsupported export file type: {path}")


def process_qc_multiverse_file(path: Path) -> list[dict]:
    payload, _ = load_payload(path)
    participant = payload.get("participant", {})
    rows = []
    for row in qc_multiverse_rows(payload):
        if not isinstance(row, dict):
            continue
        flat = {
            "file": path.name,
            "participant_id": participant.get("participantId"),
            "session_number": participant.get("session_number"),
        }
        for key, value in row.items():
            flat[key] = json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value
        rows.append(flat)
    return rows


def ordered_fieldnames(rows: list[dict]) -> list[str]:
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    return fieldnames


def build_summary_by_qc_universe(summary_rows: list[dict], qc_rows: list[dict]) -> list[dict]:
    qc_by_file: dict[str, list[dict]] = {}
    for row in qc_rows:
        file_name = row.get("file")
        if file_name:
            qc_by_file.setdefault(str(file_name), []).append(row)

    combined_rows = []
    join_keys = {"file", "participant_id", "session_number"}
    for summary in summary_rows:
        for qc in qc_by_file.get(str(summary.get("file")), []):
            combined = dict(summary)
            for key, value in qc.items():
                if key in join_keys:
                    continue
                combined[f"qc_{key}"] = value
            combined_rows.append(combined)
    return combined_rows


def collect_export_files(path: Path) -> list[Path]:
    supported = {".json", ".xlsx"}
    if path.is_file():
        return [path] if path.suffix.lower() in supported and not path.name.startswith("~$") else []
    files = []
    for suffix in supported:
        files.extend(path.glob(f"*{suffix}"))
    return sorted(p for p in files if not p.name.startswith("~$"))


def process_file(path: Path, rt_low: float, rt_high: float) -> dict:
    payload, integrity_ok = load_payload(path)
    participant = payload.get("participant", {})
    protocol = payload.get("protocol", {})
    thresholds = payload.get("outlier_thresholds", {})
    quality = payload.get("session_quality", {})
    quality_flags = payload.get("quality_flags", {})
    environment = payload.get("environment", {})
    metrics = payload.get("research_metrics", {})
    trials = payload.get("trials", {})
    qc_summary = summarize_qc_multiverse(qc_multiverse_rows(payload))

    low = thresholds.get("rt_exclude_below_ms", rt_low)
    high = thresholds.get("rt_too_slow_ms", rt_high)

    fl = summarise_task(trials.get("flanker"), low, high)
    dc = summarise_task(trials.get("dccs"), low, high)
    pc = summarise_task(trials.get("pattern-comparison"), low, high)

    # Ex-Gaussian fit on correct-trial RTs for each RT-critical task
    def correct_rts(task_trials, low=low, high=high):
        if not task_trials:
            return []
        return [t["rt"] for t in task_trials
                if isinstance(t.get("rt"), (int, float))
                and t.get("correct") == 1
                and low <= t["rt"] <= high]

    flanker_ex = fit_ex_gaussian(correct_rts(trials.get("flanker"))) or {}
    dccs_ex = fit_ex_gaussian(correct_rts(trials.get("dccs"))) or {}

    return {
        "file": path.name,
        "integrity_ok": integrity_ok,
        "participant_id": participant.get("participantId"),
        "session_number": participant.get("session_number"),
        "counterbalance_group": participant.get("counterbalance_group"),
        "consent_version": participant.get("consent_version"),
        "app_version": protocol.get("app_version", payload.get("version")),
        "protocol_version": protocol.get("protocol_version"),
        "task_version": protocol.get("task_version"),
        "scoring_version": protocol.get("scoring_version"),
        "stimulus_version": protocol.get("stimulus_version"),
        "stimulus_rendering_mode": protocol.get("stimulus_rendering_mode"),
        "age": participant.get("age"),
        "random_seed": participant.get("random_seed"),
        "viewing_distance_cm": participant.get("viewing_distance_cm"),
        "grayscale_confirmed": participant.get("grayscale_confirmed"),
        "browser": environment.get("browser"),
        "viewport_width": environment.get("viewportWidth"),
        "viewport_height": environment.get("viewportHeight"),
        "device_pixel_ratio": environment.get("devicePixelRatio"),
        "refresh_rate_hz_estimate": environment.get("refreshRateHzEstimate"),
        "tab_hidden": quality.get("visibility_hidden_count"),
        "blur_count": quality.get("blur_count"),
        "long_task_count": quality.get("long_task_count"),
        "any_quality_flag": quality_flags.get("any_quality_flag"),
        "review_recommendation": quality_flags.get("review_recommendation"),
        "review_notes": quality_flags.get("review_notes"),
        "qc_universe_count": qc_summary["qc_universe_count"],
        "qc_include_candidate_universes": qc_summary["qc_include_candidate_universes"],
        "qc_exclude_candidate_universes": qc_summary["qc_exclude_candidate_universes"],
        "qc_exclude_candidate_count": qc_summary["qc_exclude_candidate_count"],
        "low_accuracy_flag": quality_flags.get("low_accuracy_flag"),
        "fast_response_flag": quality_flags.get("fast_response_flag"),
        "slow_response_flag": quality_flags.get("slow_response_flag"),
        "many_timeouts_flag": quality_flags.get("many_timeouts_flag"),
        "focus_loss_flag": quality_flags.get("focus_loss_flag"),
        "tab_hidden_flag": quality_flags.get("tab_hidden_flag"),
        "environment_warning_flag": quality_flags.get("environment_warning_flag"),
        "environment_block_flag": quality_flags.get("environment_block_flag"),
        "flanker_n_kept_rt": fl["n_kept_rt"],
        "flanker_accuracy": fl.get("accuracy_all"),
        "flanker_mean_rt": fl.get("mean_rt_correct"),
        "flanker_inattention_flag": metrics.get("flanker_inattention_flag"),
        "flanker_congruency_effect_ms": metrics.get("flanker_congruency_effect_ms"),
        "flanker_ies_incongruent_ms": metrics.get("flanker_ies_incongruent_ms"),
        "dccs_n_kept_rt": dc["n_kept_rt"],
        "dccs_accuracy": dc.get("accuracy_all"),
        "dccs_switch_cost_ms": metrics.get("dccs_switch_cost_ms"),
        "dccs_inattention_flag": metrics.get("dccs_inattention_flag"),
        "pc_n_kept_rt": pc["n_kept_rt"],
        "pc_accuracy": pc.get("accuracy_all"),
        "pc_d_prime": metrics.get("pattern_comparison_d_prime"),
        "pc_ies_correct_ms": metrics.get("pattern_comparison_ies_correct_ms"),
        "flanker_ex_mu": flanker_ex.get("ex_mu"),
        "flanker_ex_sigma": flanker_ex.get("ex_sigma"),
        "flanker_ex_tau": flanker_ex.get("ex_tau"),
        "dccs_ex_mu": dccs_ex.get("ex_mu"),
        "dccs_ex_sigma": dccs_ex.get("ex_sigma"),
        "dccs_ex_tau": dccs_ex.get("ex_tau"),
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 1
    export_path = Path(argv[1])
    out_csv = Path(argv[2]) if len(argv) >= 3 else (
        export_path / "summary.csv" if export_path.is_dir() else export_path.with_name("summary.csv")
    )

    if not _SCIPY_AVAILABLE:
        print("[warning] scipy/numpy not installed — ex-Gaussian columns will be empty", file=sys.stderr)

    files = collect_export_files(export_path)
    if not files:
        print(f"No .json or .xlsx files found in {export_path}", file=sys.stderr)
        return 1

    rows = [process_file(p, rt_low=100, rt_high=5000) for p in files]
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {len(rows)} rows to {out_csv}")

    qc_rows: list[dict] = []
    for p in files:
        qc_rows.extend(process_qc_multiverse_file(p))
    if qc_rows:
        qc_out_csv = out_csv.parent / f"{out_csv.stem}_qc_multiverse.csv"
        fieldnames = ordered_fieldnames(qc_rows)
        with qc_out_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(qc_rows)
        print(f"Wrote {len(qc_rows)} QC multiverse rows to {qc_out_csv}")

        summary_by_qc_rows = build_summary_by_qc_universe(rows, qc_rows)
        if summary_by_qc_rows:
            summary_by_qc_out_csv = out_csv.parent / f"{out_csv.stem}_by_qc_universe.csv"
            fieldnames = ordered_fieldnames(summary_by_qc_rows)
            with summary_by_qc_out_csv.open("w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(summary_by_qc_rows)
            print(f"Wrote {len(summary_by_qc_rows)} summary-by-QC rows to {summary_by_qc_out_csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
